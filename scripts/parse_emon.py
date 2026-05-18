"""Quick triage parser for the sample EMON xlsx.

Goal: produce a one-page summary so we can ground the data contract + taxonomy
WITHOUT hauling the whole 49 MB into context. Uses openpyxl read-only mode.

Outputs to stdout in a sectioned, scannable format.
"""
from __future__ import annotations

import sys
import json
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

XLSX = Path(
    r"F:\Raptor-X-V2\rpx-core\logs\runs\Demo_ARL-Ref-U7-Dev__2026-05-15_191322_1G-1080p-H_Ultra-290K\traces\emon\emon_cyberpunk-2077_192-168-50-222_15-05-2026_200055.xlsx"
)

# A few keyword buckets to triage event-name-shaped strings.
PMU_BUCKETS = {
    "core_frontend":  ("FRONTEND", "FETCH", "DECODE", "BPU", "BR_", "IDQ", "UOPS_ISSUED", "ITLB", "ICACHE"),
    "core_backend":   ("UOPS_RETIRED", "UOPS_EXECUTED", "UOPS_DISPATCHED", "RS_", "ROB", "RAT", "EXE_ACTIVITY"),
    "core_mem":       ("L1D", "L2_", "L1I", "DTLB", "MEM_LOAD", "MEM_INST", "MEM_UOPS", "OFFCORE", "LD_BLOCKS", "STORE"),
    "core_misc":      ("CYCLES", "INST_RETIRED", "CPU_CLK", "TOPDOWN", "BR_MISP", "MACHINE_CLEARS", "INT_MISC"),
    "uncore_llc":     ("UNC_CHA", "UNC_LLC", "UNC_CBO", "UNC_HA", "LLC"),
    "uncore_imc":     ("UNC_M_", "UNC_IMC", "CAS_COUNT", "DRAM"),
    "uncore_ring":    ("UNC_RING", "UNC_R3", "UNC_R2", "UNC_QPI", "UNC_UPI"),
    "uncore_power":   ("UNC_PCU", "FREQ_", "POWER_", "PKG_", "ENERGY"),
    "graphics":       ("GT_", "GFX_", "XE_", "EU_"),
    "npu":            ("NPU_", "NCE_", "MOVIDIUS"),
    "soc":            ("SOC_", "DISPLAY_", "MEDIA_", "IPU_", "PCIE_"),
}


def bucket_for(name: str) -> str:
    if not name:
        return "blank"
    s = str(name).upper()
    for bucket, keys in PMU_BUCKETS.items():
        for k in keys:
            if k in s:
                return bucket
    return "other"


def section(title: str) -> None:
    print()
    print(f"=== {title} ===")


def main() -> None:
    if not XLSX.exists():
        print(f"NOT FOUND: {XLSX}", file=sys.stderr)
        sys.exit(2)

    print(f"File: {XLSX}")
    print(f"Size: {XLSX.stat().st_size / (1024*1024):.2f} MB")

    wb = openpyxl.load_workbook(filename=str(XLSX), read_only=True, data_only=True)
    section("SHEETS")
    for name in wb.sheetnames:
        ws = wb[name]
        # max_row/max_col are reliable on read-only after first iter, so probe
        rows = ws.max_row
        cols = ws.max_column
        print(f"  {name!r:50s}  rows={rows}  cols={cols}")

    for name in wb.sheetnames:
        ws = wb[name]
        section(f"SHEET DETAIL: {name!r}")
        # Print first 5 rows fully
        print("  -- first 5 rows --")
        rows_iter = ws.iter_rows(values_only=True)
        first_rows = []
        for i, row in enumerate(rows_iter):
            if i >= 5:
                break
            first_rows.append(row)
            preview = [str(v)[:60] if v is not None else "" for v in row[:12]]
            print(f"  row{i+1}: {preview}")

        # Try to find a header row: pick the row with the most non-null cells
        # among the first 5
        header_idx = 0
        best = -1
        for i, row in enumerate(first_rows):
            nonnull = sum(1 for v in row if v not in (None, ""))
            if nonnull > best:
                best = nonnull
                header_idx = i
        header = first_rows[header_idx] if first_rows else ()
        print(f"  -- guessed header row index: {header_idx} ({best} non-null cells) --")
        print(f"  -- header cells (truncated) --")
        for j, h in enumerate(header[:30]):
            print(f"     col{j:02d}: {h!r}")
        if len(header) > 30:
            print(f"     ... +{len(header)-30} more cols")

        # Now do a streaming pass for stats
        # We re-open because we already advanced the iterator
        ws2 = wb[name]
        row_count = 0
        bucket_counts: Counter[str] = Counter()
        sample_unmapped: list[str] = []
        first_col_counter: Counter[str] = Counter()
        second_col_counter: Counter[str] = Counter()
        col0_values = []
        col1_values = []
        for i, row in enumerate(ws2.iter_rows(values_only=True)):
            row_count += 1
            if i <= header_idx:
                continue
            # The "event name" is typically in col 0 or col 1
            v0 = row[0] if len(row) > 0 else None
            v1 = row[1] if len(row) > 1 else None
            if v0 is not None:
                first_col_counter[bucket_for(v0)] += 1
                if len(col0_values) < 8:
                    col0_values.append(v0)
            if v1 is not None:
                second_col_counter[bucket_for(v1)] += 1
                if len(col1_values) < 8:
                    col1_values.append(v1)
            # Pick whichever column has more event-shaped values
            event_name = v0 if isinstance(v0, str) and any(c.isalpha() for c in v0) else v1
            b = bucket_for(event_name)
            bucket_counts[b] += 1
            if b == "other" and len(sample_unmapped) < 10 and event_name:
                sample_unmapped.append(str(event_name))

        print(f"  -- totals -- rows={row_count}")
        print(f"  -- col0 sample values --")
        for v in col0_values:
            print(f"     {v!r}")
        print(f"  -- col1 sample values --")
        for v in col1_values:
            print(f"     {v!r}")
        print(f"  -- col0 bucket counts --")
        for b, c in sorted(first_col_counter.items(), key=lambda x: -x[1]):
            print(f"     {b:18s} {c}")
        print(f"  -- col1 bucket counts --")
        for b, c in sorted(second_col_counter.items(), key=lambda x: -x[1]):
            print(f"     {b:18s} {c}")
        print(f"  -- best-guess event-name bucket counts --")
        for b, c in sorted(bucket_counts.items(), key=lambda x: -x[1]):
            print(f"     {b:18s} {c}")
        if sample_unmapped:
            print(f"  -- sample 'other' (unmapped) event names --")
            for s in sample_unmapped:
                print(f"     {s!r}")

    wb.close()
    section("DONE")


if __name__ == "__main__":
    main()
